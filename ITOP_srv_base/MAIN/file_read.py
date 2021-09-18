
#! /usr/bin/env python
import pymysql.cursors
import sys
import os
import datetime

# -*- coding: utf-8 -*-


# Import `workbook` module from `openpyxl`
from openpyxl import Workbook

# create workbook
wb = Workbook()
# Title of the xlsx file
wb.title = "New Title"

# Create a sheet by name
sheet2=wb.create_sheet("NEW_SHEET", 0)
wb.save("srv_serial_type.xlsx")

type = []
serial = []
ip = []

sheet2.cell(1, 1, "Type")
sheet2.cell(1, 2, "SerialNumber")
sheet2.cell(1, 3, "ManagementIP")
wb.save("srv_serial_type.xlsx")

file=open('ip_type.txt')

for line in file:
    type.append(line.split()[2])
    serial.append(line.split()[0])
    ip.append(line.split()[1])
     
for i in range(2, len(type)+2):
    sheet2.cell(i, 1, type[i-2])
    sheet2.cell(i, 2, serial[i-2])
    sheet2.cell(i, 3, ip[i-2])
    wb.save("srv_serial_type.xlsx")



