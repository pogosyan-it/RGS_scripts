#! /usr/bin/env python
import pymysql.cursors  
import sys
import os
import datetime 

# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb1 = load_workbook('./Result_2020-12-04.xlsx')

# Get a sheet by name 
sheet1 = wb1.get_sheet_by_name('sheet2')

# Print the sheet title 
sheet1.title
row_max1 = sheet1.max_row
print(row_max1)

val1=sheet1.cell(row=3, column=3).value

#print(val1)

db = pymysql.connect(host="itop.rgs.ru",
                     user="itop",         # your username
                     passwd="itop",  # your password
                     db="itop")
cur = db.cursor()

for i in range(2, row_max1+1):
     end_of_war=sheet1.cell(row=i, column=3).value
     serial=sheet1.cell(row=i, column=1).value
     print(serial,'-->',end_of_war)
     query2="""UPDATE physicaldevice SET physicaldevice.end_of_warranty='%s' WHERE  physicaldevice.serialnumber='%s';""" % (end_of_war, serial)
     cur.execute(query2)
     db.commit()

