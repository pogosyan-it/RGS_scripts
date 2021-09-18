#! /usr/bin/env python
import pymysql.cursors  
import sys
import os
import datetime 
#import xlwt

# -*- coding: utf-8 -*- 


# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./export_24_04_2020.xlsx')

# Get a sheet by name 
sheet = wb.get_sheet_by_name('export')
#sheet2 = wb.get_sheet_by_name('itop_esx_list')
# Print the sheet title 
sheet.title
row_max = sheet.max_row
#print(row_max)

val1=sheet.cell(row=2, column=4).value
val1=val1+'.rgs.ru'

line = val1
str='host'+' '+line
os.system(str)
output = os.popen(str)
#print(output.read()[4])
lst = output.read().split()

print(lst[3])
# Write to the sheet of the workbook 
sheet.cell(2, 8, lst[3]) 

# Save the workbook 
wb.save("export_24_04_2020.xlsx")




db = pymysql.connect(host="itop.rgs.ru",
                     user="itop",         # your username
                     passwd="itop",  # your password
                     db="itop")
                     



cur = db.cursor()

#print(name)
#query="""Select id from view_Server where view_Server.Name='%s';""" % (name)
#print("name=",name,"serialnumber=",serialnumber,"mgmt_ip=",mgmt_ip )
#cur.execute(query)
#myresult = cur.fetchall()

for i in range(2, row_max+1):
     name=sheet.cell(row=i, column=1).value
     serialnumber=sheet.cell(row=i, column=4).value
     mgmt_ip=sheet.cell(row=i, column=7).value
     query="""Select id from view_Server where view_Server.Name='%s';""" % (name)
     cur.execute(query)
     myresult = cur.fetchall()
     for row in myresult:
         id=row[0]
         #print(id)
         query2="""UPDATE physicaldevice SET physicaldevice.serialnumber='%s' WHERE  physicaldevice.id='%s';""" % (serialnumber, id)
         query3="""UPDATE datacenterdevice SET  datacenterdevice.managementip='%s' where datacenterdevice.id='%s';""" % (mgmt_ip, id)
         cur.execute(query2)
         cur.execute(query3)
         db.commit()

