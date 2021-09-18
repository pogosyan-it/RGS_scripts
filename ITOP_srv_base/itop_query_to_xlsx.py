#! /usr/bin/env python
import pymysql.cursors  
import sys
import os
import datetime 

# -*- coding: utf-8 -*- 


# Import `workbook` module from `openpyxl`
from openpyxl import Workbook

# create workbook
wb = Workbook() 
# Title of the xlsx file
wb.title = "New Title"

# Create a sheet by name 
sheet2=wb.create_sheet("NEW_SHEET", 0)
wb.save("ESXi_list.xlsx")

db = pymysql.connect(host="itop.rgs.ru",
                     user="itop",         # your username
                     passwd="itop",  # your password
                     db="itop")
                     



cur = db.cursor()

#cur.execute('Select NOW();')

#query="""SELECT id, NAME, serialnumber, managementip, purchase_date, end_of_warranty FROM view_Server  WHERE view_Server.NAME like 'esx-%' and serialnumber=' ' or view_Server.NAME like 'esx-%' and managementip=' ';"""
query="""SELECT id, NAME, serialnumber, managementip, purchase_date, end_of_warranty FROM view_Server  
          WHERE view_Server.NAME like 'esx-%' and view_Server.`status` not in ('dismantled', 'obsolete');"""

           
cur.execute(query) 

name = []
serial = []
ip = []
date_st = []
date_end = []

sheet2.cell(1, 1, "VMHost")
sheet2.cell(1, 2, "SerialNumber")
sheet2.cell(1, 3, "ManagementIP")
sheet2.cell(1, 4, "Warranty_start")
sheet2.cell(1, 5, "Waranty_end")
wb.save("ESXi_list.xlsx")


for row in cur:
    name.append(row[1])
    serial.append(row[2])
    ip.append(row[3])
    date_st.append(row[4])
    date_end.append(row[5])


for i in range(2, len(name)+2):
     sheet2.cell(i, 1, name[i-2])
     sheet2.cell(i, 2, serial[i-2])
     sheet2.cell(i, 3, ip[i-2])
     sheet2.cell(i, 4, date_st[i-2])
     sheet2.cell(i, 5, date_end[i-2])
     wb.save("ESXi_list.xlsx")


