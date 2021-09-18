#! /usr/bin/env python
import sys
import os
import datetime 
import re
from ipaddress import ip_address
# -*- coding: utf-8 -*- 


# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./M1_SR.xlsx')

# Get a sheet by name 
sheet = wb.get_sheet_by_name('NEW_SHEET')

# Print the sheet title 
sheet.title
row_max = sheet.max_row
#print(row_max)
sheet.cell(1, 6,"HBA" )

# Save the workbook 
wb.save("M1_SR.xlsx")

for i in range(2, row_max+1):
     val1=sheet.cell(row=i,column=3).value
     #val1=val1+'.rgs.ru'
     #str='host'+' '+val1
     #str='sshpass -p "x,509SSL" ssh -o StrictHostKeyChecking=no admin@'10.200.194.36 'adapter -list' | grep FC | awk '{$1=""; print $0}
     str='sshpass -p "x,509SSL" ssh -o StrictHostKeyChecking=no admin@'+val1+' '+'\'adapter -list\' | grep FC | awk \'{$1=\"\"; print $0}\''
     print(val1)
     #os.system(str)
     #output = os.popen(str)
     #lst = output.read().split()
     #try: 
     #      ip = ip_address(lst[3])
     #      print(ip)
     #      sheet.cell(i, 7, lst[3])
     #      wb.save("export_24_04_2020.xlsx")
     #except ValueError:
     #      sheet.cell(i, 7, "NO_IP")
     #      wb.save("export_24_04_2020.xlsx") 




