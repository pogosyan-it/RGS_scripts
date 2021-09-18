#! /usr/bin/env python
import pymysql.cursors  
import sys
import os
import datetime 

# -*- coding: utf-8 -*- 


# Import `workbook` module from `openpyxl`
from openpyxl import Workbook

# create workbook
wb = Workbook() 
# Title of the xlsx file
wb.title = "New Title"

# Create a sheet by name 
sheet2=wb.create_sheet("NEW_SHEET", 0)
wb.save("M1_SR.xlsx")

db = pymysql.connect(host="itop.rgs.ru",
                     user="root",         # your username
                     passwd="qwe123qwe",  # your password
                     db="itop")
                     



cur = db.cursor()

#cur.execute('Select NOW();')

query="""SELECT view_Server.NAME, view_Server.serialnumber,view_Server.managementip, view_Server.model_name,  view_Server.rack_name
         from view_Server WHERE view_Server.location_name='M1' AND view_Server.`status` NOT IN ('obsolete', 'dismantled') 
         AND view_Server.model_name LIKE 'T%SR%';"""

           
cur.execute(query) 

name = []
serial = []
ip = []
model = []
rack = []

sheet2.cell(1, 1, "M1_SR")
sheet2.cell(1, 2, "SerialNumber")
sheet2.cell(1, 3, "ManagementIP")
sheet2.cell(1, 4, "model")
sheet2.cell(1, 5, "Rack")
wb.save("M1_SR.xlsx")


for row in cur:
    name.append(row[0])
    serial.append(row[1])
    ip.append(row[2])
    model.append(row[3])
    rack.append(row[4])

print(len(name))

for i in range(2, len(name)+2):
     sheet2.cell(i, 1, name[i-2])
     sheet2.cell(i, 2, serial[i-2])
     sheet2.cell(i, 3, ip[i-2])
     sheet2.cell(i, 4, model[i-2])
     sheet2.cell(i, 5, rack[i-2])
     wb.save("M1_SR.xlsx")


