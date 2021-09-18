#! /usr/bin/env python
import pymysql.cursors  
import sys
import os
import datetime 
#import xlwt

# -*- coding: utf-8 -*- 


# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./esx_list.xlsx')

# Get a sheet by name 
sheet = wb.get_sheet_by_name('esx_list')
sheet2 = wb.get_sheet_by_name('itop_esx_list')
# Print the sheet title 
sheet.title
row_max = sheet.max_row
#print(row_max)

val1=sheet.cell(row=2, column=4).value
val1=val1+'.rgs.ru'

line = val1
str='host'+' '+line
os.system(str)
output = os.popen(str)
#print(output.read()[4])
lst = output.read().split()

print(lst[3])
# Write to the sheet of the workbook 
sheet.cell(2, 8, lst[3]) 

# Save the workbook 
wb.save("esx_list.xlsx")

#for i in range(2, row_max+1):
     #print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=4).value)
#     val1=sheet.cell(row=i,column=4).value
#     val1=val1+'.rgs.ru'
#     str='host'+' '+val1
#     os.system(str)
#     output = os.popen(str)
#     lst = output.read().split()
#     sheet.cell(i, 8, lst[3])
#     wb.save("esx_list.xlsx")




db = pymysql.connect(host="itop.rgs.ru",
                     user="itop",         # your username
                     passwd="itop",  # your password
                     db="itop")
                     



cur = db.cursor()

#cur.execute('Select NOW();')

query="""SELECT id, NAME, serialnumber, managementip, purchase_date, end_of_warranty FROM view_Server 
         WHERE view_Server.NAME like 'esx-%' and serialnumber=' ' or view_Server.NAME like 'esx-%' and managementip=' '"""

query2=""" Select id from view_Server where view_Server.Name="%s" """
           
cur.execute(query) 

name = []
serial = []
ip = []
date_st = []
date_end = []

for row in cur:
    #l=row[0].split('.') 
    #name=row[0].split(',')
    #print(row[0],row[1],row[2],row[3],row[4],row[5]) 
    name.append(row[1])
    serial.append(row[2])
    ip.append(row[3])
    date_st.append(row[4])
    date_end.append(row[5])


for i in range(1, len(name)+1):
     #print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=4).value)
#     val1=sheet.cell(row=i,column=4).value
#     val1=val1+'.rgs.ru'
#     str='host'+' '+val1
#     os.system(str)
#     output = os.popen(str)
#     lst = output.read().split()
     sheet2.cell(i, 1, name[i-1])
     sheet2.cell(i, 2, serial[i-1])
     sheet2.cell(i, 3, ip[i-1])
     sheet2.cell(i, 4, date_st[i-1])
     sheet2.cell(i, 5, date_end[i-1])
     #print(name[i])1111
     wb.save("esx_list.xlsx")


#print(len(row[0]))

#for i in range(len(name)):
#    cur.execute(query2,(name[i]))  
#    myresult2 = cur.fetchall()
#    for row2 in myresult2:
#        print(row2)

#def isint(s):
#    try:
#        int(s)
#        return True
#    except ValueError:
#        return False


