#! /usr/bin/env python
import sys
import os
import datetime 
import re
from ipaddress import ip_address
# -*- coding: utf-8 -*- 


# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./export_24_04_2020.xlsx')

# Get a sheet by name 
sheet = wb.get_sheet_by_name('export')

# Print the sheet title 
sheet.title
row_max = sheet.max_row
#print(row_max)
sheet.cell(1, 7,"ManagementIP" )

# Save the workbook 
wb.save("esx_list.xlsx")

for i in range(2, row_max+1):
     val1=sheet.cell(row=i,column=4).value
     val1=val1+'.rgs.ru'
     str='host'+' '+val1
     os.system(str)
     output = os.popen(str)
     lst = output.read().split()
     try: 
           ip = ip_address(lst[3])
           print(ip)
           sheet.cell(i, 7, lst[3])
           wb.save("export_24_04_2020.xlsx")
     except ValueError:
           sheet.cell(i, 7, "NO_IP")
           wb.save("export_24_04_2020.xlsx") 




